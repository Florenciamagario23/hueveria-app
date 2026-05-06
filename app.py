from flask import Flask, render_template, request, redirect, session, send_file
from werkzeug.security import generate_password_hash
import pandas as pd
import io
import urllib.parse
from datetime import datetime
from openpyxl import Workbook
from flask import send_file
import io
from datetime import datetime
import psycopg2
import psycopg2.extras
import os
from dotenv import load_dotenv
load_dotenv()


url = os.environ.get("DATABASE_URL")

app = Flask(__name__)
app.secret_key = "clave_secreta"


#CONEXION
def conectar():
    url = os.environ.get("DATABASE_URL")

    if not url:
        raise Exception("❌ DATABASE_URL no está configurada")

    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)

    return psycopg2.connect(
        url,
        cursor_factory=psycopg2.extras.RealDictCursor
    )

# 🧱 CREAR TABLAS
def crear_tablas():
    conn = conectar()
    cursor = conn.cursor()

    # 👤 USUARIOS
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id SERIAL PRIMARY KEY,
        usuario TEXT UNIQUE,
        password TEXT
    )
    """)

    # 📦 PRODUCTOS
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS productos (
        id SERIAL PRIMARY KEY,
        nombre TEXT,
        stock_inicial INTEGER,
        stock_actual INTEGER,
        precio NUMERIC
    )
    """)

    # 💰 VENTAS
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS ventas (
        id SERIAL PRIMARY KEY,
        fecha TIMESTAMP,
        producto_id INTEGER,
        cantidad INTEGER,
        total NUMERIC,
        metodo_pago TEXT,
        eliminado INTEGER DEFAULT 0
    )
    """)

    # 💸 GASTOS
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS gastos (
        id SERIAL PRIMARY KEY,
        fecha TIMESTAMP,
        descripcion TEXT,
        monto NUMERIC,
        eliminado INTEGER DEFAULT 0
    )
    """)

    conn.commit()
    conn.close()

# 👤 USUARIO
def crear_usuario():
    conn = conectar()
    cursor = conn.cursor()

    usuarios = ["Micaela", "Magali", "Francisco"]
    password = generate_password_hash("Familia26@")

    for u in usuarios:
        cursor.execute("SELECT * FROM usuarios WHERE usuario = %s", (u,))
        if not cursor.fetchone():
            cursor.execute(
                "INSERT INTO usuarios (usuario, password) VALUES (%s, %s)",
                (u, password)
            )

    conn.commit()
    conn.close()

def cargar_productos_base():
    conn = conectar()
    cursor = conn.cursor()

    # 🔥 VERIFICAR SI YA HAY PRODUCTOS
    cursor.execute("SELECT COUNT(*) AS total FROM productos")
    cantidad = cursor.fetchone()["total"]

    if cantidad > 0:
        conn.close()
        return  # 👈 NO hace nada si ya hay datos

    productos = [
        ("MAPLE DE HUEVO CHICO COLOR/BLANCO", 5000),
        ("MAPLE DE HUEVO MEDIANO COLOR/BLANCO", 6000),
        ("MAPLE DE HUEVO MEDIANITO COLOR/BLANCO", 5400),
        ("MAPLE DE HUEVO GRANDE COLOR/BLANCO", 7200)
    ]

    for nombre, precio in productos:
        cursor.execute("""
            INSERT INTO productos (nombre, stock_inicial, stock_actual, precio)
            VALUES (%s, 450, 450, %s)
        """, (nombre, precio))

    conn.commit()
    conn.close()

def arreglar_db():
    conn = conectar()
    cursor = conn.cursor()

    try:
        cursor.execute("ALTER TABLE ventas ADD COLUMN metodo_pago TEXT")
    except:
        pass

    try:
        cursor.execute("ALTER TABLE ventas ADD COLUMN eliminado INTEGER DEFAULT 0")
    except:
        pass

    try:
        cursor.execute("ALTER TABLE gastos ADD COLUMN eliminado INTEGER DEFAULT 0")
    except:
        pass

    try:
        cursor.execute("ALTER TABLE productos ADD COLUMN fecha_stock TEXT")
    except:
        pass

    conn.commit()
    conn.close()


def actualizar_stock_diario():
    conn = conectar()
    cursor = conn.cursor()

    hoy = datetime.now().strftime("%Y-%m-%d")

    cursor.execute("SELECT id, stock_actual, fecha_stock FROM productos")
    productos = cursor.fetchall()

    for p in productos:
        if p["fecha_stock"] != hoy:
            cursor.execute("""
                UPDATE productos
                SET stock_inicial = stock_actual,
                    fecha_stock = %s
                WHERE id = %s
            """, (hoy, p["id"]))

    conn.commit()
    conn.close()


# 🔐 LOGIN
from werkzeug.security import check_password_hash

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"].strip()
        password = request.form["password"].strip()

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM usuarios WHERE usuario = %s", (usuario,))
        user = cursor.fetchone()

        if user and check_password_hash(user["password"], password):
            session["usuario"] = usuario
            return redirect("/dashboard")

        return render_template("login.html", error="Datos incorrectos")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# 🏠 HOME
@app.route("/")
def home():
    carrito = session.get("carrito", [])
    cantidad_carrito = sum(item["cantidad"] for item in carrito)
    return render_template("home.html", cantidad_carrito=cantidad_carrito)

# 🛒 AGREGAR CARRITO
@app.route("/agregar", methods=["POST"])
def agregar():
    nombre = request.form.get("producto")
    precio = int(request.form.get("precio"))

    carrito = session.get("carrito", [])

    for item in carrito:
        if item["nombre"] == nombre:
            item["cantidad"] += 1
            session["carrito"] = carrito
            return redirect("/")

    carrito.append({
        "nombre": nombre,
        "precio": precio,
        "cantidad": 1
    })

    session["carrito"] = carrito
    return redirect("/")

# 🛒 VER CARRITO
@app.route("/carrito")
def ver_carrito():
    carrito = session.get("carrito", [])
    total = sum(item["precio"] * item["cantidad"] for item in carrito)
    return render_template("carrito.html", carrito=carrito, total=total)

# ➕➖
@app.route("/sumar/<nombre>")
def sumar(nombre):
    for item in session["carrito"]:
        if item["nombre"] == nombre:
            item["cantidad"] += 1
    session.modified = True
    return redirect("/carrito")

@app.route("/restar/<nombre>")
def restar(nombre):
    for item in session["carrito"]:
        if item["nombre"] == nombre:
            item["cantidad"] -= 1
            if item["cantidad"] <= 0:
                session["carrito"].remove(item)
    session.modified = True
    return redirect("/carrito")

@app.route("/vaciar")
def vaciar():
    session.pop("carrito", None)
    return redirect("/carrito")

# 📲 WHATSAPP
@app.route("/enviar")
def enviar():
    carrito = session.get("carrito", [])
    mensaje = "Hola, quiero pedir:\n\n"
    total = 0

    for item in carrito:
        subtotal = item["precio"] * item["cantidad"]
        total += subtotal
        mensaje += f"{item['nombre']} x{item['cantidad']} = ${subtotal}\n"

    mensaje += f"\nTOTAL: ${total}"

    mensaje_codificado = urllib.parse.quote(mensaje)
    return redirect(f"https://wa.me/5493564593629?text={mensaje_codificado}")

@app.route("/dashboard")
def dashboard():
    if "usuario" not in session:
        return redirect("/login")

    conn = conectar()
    cursor = conn.cursor()

    # 🔥 Totales
    cursor.execute("SELECT COALESCE(SUM(total),0) FROM ventas WHERE eliminado = 0")
    ventas_total = list(cursor.fetchone().values())[0]

    cursor.execute("""
SELECT COALESCE(SUM(monto),0) AS total
FROM gastos
WHERE eliminado = 0
""")
    gastos_total = cursor.fetchone()["total"]

    ganancia = ventas_total - gastos_total

    # 🔥 VENTAS
    cursor.execute("""
        SELECT v.*, p.nombre
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE v.eliminado = 0
        ORDER BY v.fecha DESC
    """)
    ventas_todas = cursor.fetchall()

    ventas_ultimas = ventas_todas[:3]  # 👈 SOLO 3

    # 🔥 GASTOS
    cursor.execute("""
        SELECT *
        FROM gastos
        WHERE eliminado = 0
        ORDER BY fecha DESC
    """)
    gastos_todos = cursor.fetchall()

    gastos_ultimos = gastos_todos[:3]  # 👈 SOLO 3

    # 🔥 PRODUCTOS
    cursor.execute("SELECT * FROM productos")
    productos = cursor.fetchall()

    conn.close()

    ventas_total_cantidad = len(ventas_todas)
    gastos_total_cantidad = len(gastos_todos)

    return render_template(
    "ventas.html",
    ventas_ultimas=ventas_ultimas,
    ventas_todas=ventas_todas,
    gastos_ultimos=gastos_ultimos,
    gastos_todos=gastos_todos,
    productos=productos,
    ventas_total=ventas_total,
    gastos_total=gastos_total,
    ganancia=ganancia,
    ventas_total_cantidad=ventas_total_cantidad,
    gastos_total_cantidad=gastos_total_cantidad
)

@app.route("/eliminar_venta/<int:id>", methods=["POST"])
def eliminar_venta(id):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("UPDATE ventas SET eliminado = 1 WHERE id = %s", (id,))
    
    conn.commit()
    conn.close()

    return redirect("/dashboard")

@app.route("/eliminar_gasto/<int:id>", methods=["POST"])
def eliminar_gasto(id):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("UPDATE gastos SET eliminado = 1 WHERE id = %s", (id,))
    
    conn.commit()
    conn.close()

    return redirect("/dashboard")

@app.route("/historial")
def historial():
    conn = conectar()
    cursor = conn.cursor()

    # 🔴 VENTAS ELIMINADAS
    cursor.execute("""
    SELECT v.id, 'VENTA', p.nombre, v.fecha
    FROM ventas v
    JOIN productos p ON v.producto_id = p.id
    WHERE v.eliminado = 1
    """)
    ventas = cursor.fetchall()

    # 🔵 GASTOS ELIMINADOS
    cursor.execute("""
    SELECT id, 'GASTO', descripcion, fecha
    FROM gastos
    WHERE eliminado = 1
    """)
    gastos = cursor.fetchall()

    conn.close()

    historial = ventas + gastos

    # ordenar por fecha DESC
    historial.sort(key=lambda x: x[3], reverse=True)

    return render_template("historial.html", historial=historial)

@app.route("/restaurar_historial/<int:id>", methods=["POST"])
def restaurar_historial(id):
    conn = conectar()
    cursor = conn.cursor()

    # intenta restaurar en ventas
    cursor.execute("UPDATE ventas SET eliminado = 0 WHERE id = %s", (id,))
    
    # intenta restaurar en gastos
    cursor.execute("UPDATE gastos SET eliminado = 0 WHERE id = %s", (id,))

    conn.commit()
    conn.close()

    return redirect("/historial")

@app.route("/eliminar_historial/<int:id>", methods=["POST"])
def eliminar_historial(id):
    conn = conectar()
    cursor = conn.cursor()

    # borrar de ventas si existe
    cursor.execute("DELETE FROM ventas WHERE id = %s", (id,))
    
    # borrar de gastos si existe
    cursor.execute("DELETE FROM gastos WHERE id = %s", (id,))

    conn.commit()
    conn.close()

    return redirect("/historial")

# 💰 VENTA
@app.route("/agregar_venta", methods=["POST"])
def agregar_venta():
    producto_id = int(request.form["producto_id"])

    try:
        cantidad = int(request.form["cantidad"])
    except:
        cantidad = 0

    metodo_pago = request.form["metodo_pago"]

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT precio, stock_actual FROM productos WHERE id = %s",
        (producto_id,)
    )
    producto = cursor.fetchone()

    if not producto:
        return "Producto no encontrado"

    precio = producto["precio"]
    stock = producto["stock_actual"]

    if cantidad > stock:
        return "Sin stock"

    total = precio * cantidad

    fecha = request.form.get("fecha")

    if not fecha:
     fecha = datetime.now()

    cursor.execute("""
    INSERT INTO ventas (fecha, producto_id, cantidad, total, metodo_pago)
    VALUES (%s, %s, %s, %s, %s)
""", (fecha, producto_id, cantidad, total, metodo_pago))

    cursor.execute("""
        UPDATE productos
        SET stock_actual = stock_actual - %s
        WHERE id = %s
    """, (cantidad, producto_id))

    conn.commit()
    conn.close()

    return redirect("/dashboard")

# 💸 GASTO
@app.route("/agregar_gasto", methods=["POST"])
def agregar_gasto():
    descripcion = request.form["descripcion"]
    monto = float(request.form["monto"])

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO gastos (descripcion, monto, fecha)
        VALUES (%s, %s, NOW())
    """, (descripcion, monto))

    conn.commit()
    conn.close()

    return redirect("/dashboard")

@app.route("/agregar_producto", methods=["POST"])
def agregar_producto():
    conn = conectar()
    cursor = conn.cursor()

    nombre = request.form["nombre"]
    precio = int(request.form["precio"])
    stock = int(request.form["stock"])

    # evitar duplicados
    cursor.execute("SELECT * FROM productos WHERE nombre = %s", (nombre,))
    existe = cursor.fetchone()

    if not existe:
        cursor.execute("""
            INSERT INTO productos (nombre, stock_inicial, stock_actual, precio)
            VALUES (%s, %s, %s, %s)
        """, (nombre, stock, stock, precio))

    conn.commit()
    conn.close()

    return redirect("/dashboard")

@app.route("/actualizar_stock/<int:id>", methods=["POST"])
def actualizar_stock(id):
    stock = request.form["stock"]
    precio = request.form["precio"]

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE productos
        SET stock_actual = %s, precio = %s
        WHERE id = %s
    """, (stock, precio, id))

    conn.commit()
    conn.close()

    return redirect("/dashboard")

@app.route("/eliminar_producto/<int:id>", methods=["POST"])
def eliminar_producto(id):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM productos WHERE id = %s", (id,))

    conn.commit()
    conn.close()

    return redirect("/dashboard")

def arreglar_tabla_productos():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        ALTER TABLE productos 
        ADD COLUMN IF NOT EXISTS fecha_stock TEXT
    """)

    cursor.execute("""
        ALTER TABLE productos 
        ADD COLUMN IF NOT EXISTS stock_inicial INTEGER DEFAULT 0
    """)

    cursor.execute("""
        ALTER TABLE productos 
        ADD COLUMN IF NOT EXISTS stock_actual INTEGER DEFAULT 0
    """)

    conn.commit()
    conn.close()

    from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from flask import send_file
import io
from datetime import datetime

def ajustar_columnas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 3


@app.route("/exportar_excel")
def exportar_excel():
    conn = conectar()
    cursor = conn.cursor()

    hoy = datetime.now().strftime("%Y-%m-%d")

    # 📊 VENTAS
    cursor.execute("""
    SELECT 
        v.fecha,
        p.nombre,
        v.cantidad,
        v.total,
        v.metodo_pago
    FROM ventas v
    JOIN productos p ON v.producto_id = p.id
    WHERE v.eliminado = 0
    ORDER BY v.fecha DESC
""")
    ventas = cursor.fetchall()

    # 💸 GASTOS
    cursor.execute("""
        SELECT descripcion, monto
        FROM gastos
        WHERE DATE(fecha) = %s
    """, (hoy,))
    gastos = cursor.fetchall()

    conn.close()

    wb = Workbook()

    # =============================
    # 🟢 HOJA 1: RESUMEN
    # =============================
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = f"Resumen del día {hoy}"
    ws["A1"].font = Font(size=14, bold=True)

    total_ventas = sum(v["total"] for v in ventas)
    total_gastos = sum(g["monto"] for g in gastos)
    ganancia = total_ventas - total_gastos

    ws["A3"] = "Ventas"
    ws["B3"] = total_ventas
    ws["A4"] = "Gastos"
    ws["B4"] = total_gastos
    ws["A5"] = "Ganancia"
    ws["B5"] = ganancia

    # Colores
    ws["B3"].fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    ws["B4"].fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    ws["B5"].fill = PatternFill(start_color="BDD7EE", fill_type="solid")

    # Formato $
    for c in ["B3", "B4", "B5"]:
        ws[c].number_format = '"$"#,##0'

    ajustar_columnas(ws)

    # =============================
    # 🟢 HOJA 2: VENTAS
    # =============================
    ws_ventas = wb.create_sheet("Ventas")

    ws_ventas.append(["Fecha", "Producto", "Cantidad", "Total", "Pago"])

    for v in ventas:
     ws_ventas.append([
        v["fecha"],
        v["nombre"],
        v["cantidad"],
        v["total"],
        v["metodo_pago"]
    ])

    # Estilo encabezado
    header_fill = PatternFill(start_color="333333", fill_type="solid")

    for cell in ws_ventas[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Formato $
    for row in ws_ventas.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '"$"#,##0'

    ajustar_columnas(ws_ventas)

    # =============================
    # 🔴 HOJA 3: GASTOS
    # =============================
    ws_gastos = wb.create_sheet("Gastos")

    ws_gastos.append(["Descripción", "Monto"])

    for g in gastos:
        ws_gastos.append([g["descripcion"], g["monto"]])

    for cell in ws_gastos[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in ws_gastos.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '"$"#,##0'

    ajustar_columnas(ws_gastos)

    # =============================
    # 📊 HOJA 4: GRÁFICO
    # =============================
    ws_chart = wb.create_sheet("Gráfico")

    ws_chart.append(["Producto", "Total"])

    for v in ventas:
        ws_chart.append([v["nombre"], v["total"]])

    chart = BarChart()
    chart.title = "Ventas por producto"
    chart.y_axis.title = "Total"
    chart.x_axis.title = "Producto"

    data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(ventas)+1)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(ventas)+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws_chart.add_chart(chart, "D2")

    ajustar_columnas(ws_chart)

    # =============================
    # 📥 EXPORTAR
    # =============================
    archivo = io.BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"reporte_{hoy}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 🚀 INIT
crear_tablas()
arreglar_tabla_productos()  # 👈 SIEMPRE antes de usar productos
crear_usuario()
arreglar_db()
cargar_productos_base()
actualizar_stock_diario()


if __name__ == "__main__":
    app.run(debug=True)